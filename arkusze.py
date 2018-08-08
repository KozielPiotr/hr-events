import openpyxl
from openpyxl.utils import column_index_from_string
from itertools import groupby
from operator import itemgetter
from os import system, name


def welcome_screen():
    system("cls" if name == "nt" else "clear")
    print(80*"=")
    print("||" + " "*24 + "Zestawienia grafikowe v.1.0" + " "*25 + "||")
    print("||" + " " * 28 + "autor: Piotr Kozieł" + " " * 29 + "||")
    print(80*"=")


# Describes to user "to do" and "not to do" stuff
def instructions():
    welcome_screen()
    print("\nProgram poprosi Cię o wprowadzenie ilości sklepów, dla których chcesz utworzyć zestawienie.")
    print("Następnie z grafików, które wprowadzisz wyciągnie potrzebne dane i ewentualnie poprosi Cię o wpisanie dodatkowych informacji.")
    print("\n1. Grafik musi być w formacie xlsx, a nie xls - inaczej nie działa. Zmiana rozszerzenia nic nie da, trzeba zapisać plik w tym formacie.")
    print("2. miesiąc musi być wpisany słownie, inaczej nie zadziała.")
    print("3. nazwę grafiku najlepiej skopiować i wkleić, jeśli będzie się różnić od nazwy pliku, to kaplica.")
    print("\n!!!\nZłamanie powyższych punktów spowoduje wysypanie się programu, jeszcze nie wprowadziłem zabezpieczeń ;)\n!!!")
    print("\nMam nadzieję, że program się przyda ;)   A teraz [enter] i do dzieła!")


# returns a month it's number
def num_month(month):
    if month == "Styczeń":
        n_m = 1
    elif month == "Luty":
        n_m = 2
    elif month == "Marzec":
        n_m = 3
    elif month == "Kwiecień":
        n_m = 4
    elif month == "Maj":
        n_m = 5
    elif month == "Czerwiec":
        n_m = 6
    elif month == "Lipiec":
        n_m = 7
    elif month == "Sierpień":
        n_m = 8
    elif month == "Wrzesień":
        n_m = 9
    elif month == "Październik":
        n_m = 10
    elif month == "Listopad":
        n_m = 11
    elif month == "Grudzień":
        n_m = 12
    else:
        return False
    return n_m


# looks for shop's name and returns it's cell as a variable
def lf_shop(sheet, our_shops):
    shop = {}
    for i in range(int(sheet.max_row + 1)):
        for j in range(int(sheet.max_column + 1)):
            if sheet.cell(row=i+1, column=j+1).value in our_shops:
                shop = sheet.cell(row=i+1, column=j+1)
    return shop


# returns dictionary: [worker's_number: worker's_name]
def list_of_workers(sheet, shop_cell):
    workers = {}
    for i in range(4, int(sheet.max_column)):
        if type(sheet.cell(row=shop_cell.row, column=i).value) == str:
            name = (sheet.cell(row=shop_cell.row, column=i)).value
            surname = (sheet.cell(row=shop_cell.row+1, column=i)).value
            workers[i] = (name + " " + surname) # key in dictionary is a number of column for each worker
    return workers


# counts sum of hours worked in month by employee
def count_working_hours(w_h, current_day, active_cell, sheet):
    try:
        if type(current_day.value) == int:
            w_h += sheet.cell(row=current_day.row, column=column_index_from_string(active_cell.column) - 1).value
        return w_h
    except:
        c = sheet.cell(row=current_day.row, column=column_index_from_string(active_cell.column) - 1)
        w_h += int(input("Błąd w grafiku. Sprawdź komórkę %s%s i wpisz prawidlowa liczbe godzin pracy tego dnia:\n" % (c.row, c.column)))
        return w_h


# looks for events
def event_check(e_lst, current_day, active_cell, event, sheet, i_obj):
    if type(current_day.value) == int:
        active_cell = sheet.cell(row=current_day.row, column=int(i_obj) + 3)
        if active_cell.value == event:
            e_lst.append(current_day.value)
    return e_lst


# prints vacation leave days
def print_vl(lst, month_number, y):
    ranges = []
    for k, g in groupby(enumerate(lst), lambda x: x[0] - x[1]):
        group = (map(itemgetter(1), g))
        group = list(map(int, group))
        ranges.append((group[0], group[-1]))
        vl = (ranges[-1])
        f_day = vl[0]
        l_day = vl[1]
        if f_day != l_day:
            outfile = (open("zestawienie.txt", "a"))
            outfile.write("- Urlop wypoczynkowy w dniach: %02d-%02d.%02d.%04d\n" % (f_day, l_day, month_number, y.value))
            outfile.close()
        else:
            outfile = (open("zestawienie.txt", "a"))
            outfile.write("- Urlop wypoczynkowy w dniu: %02d.%02d.%04d\n" % (f_day, month_number, y.value))
            outfile.close()


# prints leave on request days
def print_lor(lst, month_number, y, sheet, i_obj, coord_cell):
    for day in lst:
        begin_hour = sheet.cell(row=coord_cell.row+2+day, column=i_obj).value
        end_hour = sheet.cell(row=coord_cell.row+2+day, column=i_obj+1).value
        outfile = (open("zestawienie.txt", "a"))
        outfile.write("- urlop na żądanie w dniu %02d.%02d.%04d, gdzie miał pracować %02d:00-%02d:00\n" % (day, month_number, y.value, begin_hour, end_hour))
        outfile.close()


def print_l4(wkr, lst, month_number, y, shop):
    ranges = []
    for k, g in groupby(enumerate(lst), lambda x: x[0] - x[1]):
        group = (map(itemgetter(1), g))
        group = list(map(int, group))
        ranges.append((group[0], group[-1]))
        L4 = (ranges[-1])
        f_day = L4[0]
        l_day = L4[1]
        if f_day != l_day:
            welcome_screen()
            L4_no = input("%s: %s: podaj nr L4 z dni %02d-%02d.%02d.%04d" % (shop, wkr, f_day, l_day, month_number, y.value))
            outfile = (open("zestawienie.txt", "a"))
            outfile.write("- Zwolnienie L4 w dniach: %02d-%02d.%02d.%04d. Numer zwolnienia:  %s\n" % (f_day, l_day, month_number, y.value, L4_no.upper()))
            outfile.close()
        else:
            welcome_screen()
            L4_no = input("%s: %s: podaj nr L4 z dnia %02d.%02d.%04d" % (shop, wkr, f_day, month_number, y.value))
            outfile = (open("zestawienie.txt", "a"))
            outfile.write("- Zwolnienie L4 w dniu: %02d.%02d.%04d. Numer zwolnienia:  %s\n" % (f_day, month_number, y.value, L4_no.upper()))
            outfile.close()


# prints special leave days
def print_sl(lst, month_number, y):
    ranges = []
    for k, g in groupby(enumerate(lst), lambda x: x[0] - x[1]):
        group = (map(itemgetter(1), g))
        group = list(map(int, group))
        ranges.append((group[0], group[-1]))
        sl = (ranges[-1])
        f_day = sl[0]
        l_day = sl[1]
        if f_day != l_day:
            outfile = (open("zestawienie.txt", "a"))
            outfile.write("- Urlop okolicznościowy w dniach: %02d-%02d.%02d.%04d\n" % (f_day, l_day, month_number, y.value))
            outfile.close()
        else:
            outfile = (open("zestawienie.txt", "a"))
            outfile.write("- Urlop okolicznościowy w dniu: %02d.%02d.%04d\n" % (f_day, month_number, y.value))
            outfile.close()


# prints father's leave days
def print_fl(lst, month_number, y):
    ranges = []
    for k, g in groupby(enumerate(lst), lambda x: x[0] - x[1]):
        group = (map(itemgetter(1), g))
        group = list(map(int, group))
        ranges.append((group[0], group[-1]))
        fl = (ranges[-1])
        f_day = fl[0]
        l_day = fl[1]
        if f_day != l_day:
            outfile = (open("zestawienie.txt", "a"))
            outfile.write("- Urlop ojcowski w dniach: %02d-%02d.%02d.%04d\n" % (f_day, l_day, month_number, y.value))
            outfile.close()
        else:
            outfile = (open("zestawienie.txt", "a"))
            outfile.write("- Urlop ojcowski w dniu: %02d.%02d.%04d\n" % (f_day, month_number, y.value))
            outfile.close()


# gets everything together and prints to file
def create(m, file):
    n_m = num_month(m)

    # list of our shops
    shops = ["DBC", "DKD", "DGS", "DGK", "DO", "DPP", "DPA", "DSP", "DWW", "DWB", "DLM", "DWT"]

    # opening work schedule file
    wb = openpyxl.load_workbook("%s.xlsx" % file, data_only=True)
    sh = wb["Grafik_" + m]

    # looking for shop's name and setting it's code to variable
    shop = lf_shop(sh, shops) # variable set to cell object

    # making list of workers
    workers = list_of_workers(sh, shop)

    shop_n_row = shop.row
    shop_n_col = column_index_from_string(shop.column)
    # looking for year
    year = sh.cell(row=shop_n_row+2, column=shop_n_col)

    outfile = open("zestawienie.txt", "a")
    outfile.write("\n\n\n" + shop.value.upper()  + ":\n")
    outfile.close()


    # itrates through workers
    for worker in workers:
        cur_worker = str(workers[worker])
        worker_hours = 0
        outfile = open("zestawienie.txt", "a")
        outfile.write("\n%s:\n" % cur_worker)
        outfile.close()
        a_cell = sh.cell(row=shop_n_row+3, column=int(worker)+3)
        lst_vl = []
        lst_lor = []
        lst_l4 = []
        lst_sl = []
        lst_fl = []
        for i in range(a_cell.row, sh.max_row):
            c_day = sh.cell(row=i, column=shop_n_col+1)
            worker_hours = count_working_hours(worker_hours, c_day, a_cell, sh)

            # looks for event
            vl = "UW"
            vl_days = event_check(lst_vl, c_day, a_cell, vl, sh, worker)

            lor = "UNŻ"
            lor_days = event_check(lst_lor, c_day, a_cell, lor, sh, worker)

            l4 = "L4"
            l4_days = event_check(lst_l4, c_day, a_cell, l4, sh, worker)

            sl = "UO"
            sl_days = event_check(lst_sl, c_day, a_cell, sl, sh, worker)

            fl = "UOJ"
            fl_days = event_check(lst_fl, c_day, a_cell, fl, sh, worker)

        print_vl(vl_days, n_m, year)
        print_lor(lor_days, n_m, year, sh, worker, shop)
        print_l4(workers[worker], l4_days, n_m, year, shop.value)
        print_sl(sl_days, n_m, year)
        print_fl(fl_days, n_m, year)

        outfile = (open("zestawienie.txt", "a"))
        outfile.write("- %s w miesiącu %s przepracuje %s godzin\n" % (workers[worker], m.lower(), worker_hours))
        outfile.close()


def zestawienie():
    instructions()
    input()
    welcome_screen()
    n_o_s_h = []
    number_of_shops = int(input(3*"\n" + 22* " " + "Dla ilu sklepów ma być zestawienie?"))
    for i in range(0, number_of_shops):
        welcome_screen()
        n_o_s_h.append(input(6 * " " + "Nazwa pliku z %s grafikiem (skopiuj nazwe i wklej bez rozszerzenia):\n" % str(i+1)))
    month = (input(3*"\n" + 27*" " + "Wprowadz słownie miesiąc: \n" + 36*" ")).title()

    welcome_screen()
    outfile = open("zestawienie.txt", "w")
    outfile.write("Zestawienie dla miesiąca: %s" % month)
    outfile.close()
    for shop in n_o_s_h:
        create(month, shop)
        welcome_screen()
    print("\n"*3 + " " * 31 + "Zestawienie gotowe\n" + " " * 24 + 'Wyniki w pliku "zestawienie.txt"')



zestawienie()
input()